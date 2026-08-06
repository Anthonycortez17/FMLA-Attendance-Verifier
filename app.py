import io
import re
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st

# Page Configuration
st.set_page_config(
    page_title="HRIS Leave Claim Automation | Benefits Analytics",
    page_icon="🔴",
    layout="wide"
)

# ==============================================================================
# COSTCO LOGO & DYNAMIC DARK/LIGHT ADAPTIVE CSS
# ==============================================================================
COSTCO_LOGO_URL = "https://upload.wikimedia.org/wikipedia/commons/5/59/Costco_Wholesale_logo.svg"

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    /* Global Typography */
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif !important;
    }

    /* Executive Costco Header */
    .costco-header-card {
        background: linear-gradient(135deg, #002B49 0%, #00437A 60%, #005DAA 100%);
        padding: 24px 30px;
        border-radius: 16px;
        border-bottom: 5px solid #E31837;
        box-shadow: 0 8px 24px rgba(0, 43, 73, 0.25);
        margin-bottom: 25px;
        display: flex;
        align-items: center;
        gap: 24px;
    }

    .costco-logo-badge {
        background-color: #FFFFFF;
        padding: 10px 18px;
        border-radius: 12px;
        display: flex;
        align-items: center;
        justify-content: center;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.2);
        min-width: 150px;
    }

    .costco-title-text {
        color: #FFFFFF !important;
        font-size: 2.1rem;
        font-weight: 800;
        margin: 0;
        letter-spacing: -0.5px;
        line-height: 1.2;
    }

    .costco-subtitle-text {
        color: #CBD5E1 !important;
        font-size: 1.02rem;
        font-weight: 500;
        margin-top: 6px;
    }

    /* Adaptive Card Containers (Supports both Light and Dark themes) */
    .card-container {
        background-color: rgba(0, 93, 170, 0.05);
        border: 1px solid rgba(0, 93, 170, 0.2);
        border-left: 6px solid #005DAA;
        padding: 20px 24px;
        border-radius: 12px;
        margin-bottom: 20px;
        backdrop-filter: blur(8px);
    }

    .card-header-title {
        font-size: 1.15rem;
        font-weight: 700;
        margin-bottom: 6px;
        display: flex;
        align-items: center;
        gap: 8px;
    }

    /* Costco Red Primary Button */
    .stButton>button {
        width: 100%;
        background: linear-gradient(135deg, #E31837 0%, #C0132D 100%) !important;
        color: #FFFFFF !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        border-radius: 10px !important;
        border: none !important;
        height: 3.4em !important;
        box-shadow: 0 4px 14px rgba(227, 24, 55, 0.35) !important;
        transition: all 0.25s ease-in-out !important;
    }

    .stButton>button:hover {
        background: linear-gradient(135deg, #C0132D 0%, #9E0C22 100%) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(227, 24, 55, 0.5) !important;
    }

    /* Input Fields Styling */
    .stTextInput input, .stTextArea textarea {
        border-radius: 10px !important;
        font-size: 0.98rem !important;
    }

    /* Expander Container Adaptive Styling */
    [data-testid="stExpander"] {
        border-radius: 12px !important;
        border: 1px solid rgba(0, 93, 170, 0.2) !important;
    }
    </style>
""", unsafe_allow_html=True)

# ==============================================================================
# HEADER BANNER & COSTCO LOGO
# ==============================================================================
banner_col1, banner_col2 = st.columns([1, 5])

with banner_col1:
    st.image(COSTCO_LOGO_URL, use_container_width=True)

with banner_col2:
    st.markdown("""
        <div class="costco-header-card">
            <div>
                <div class="costco-title-text">HRIS FMLA & Leave Claim Automation</div>
                <div class="costco-subtitle-text">Automated Unum Request & SAP Time Verification Platform | Benefits Analytics</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

# ==============================================================================
# SESSION USER SETTINGS (Dynamic Analyst Name & Title)
# ==============================================================================
st.markdown("""
    <div class="card-container">
        <div class="card-header-title">👤 Analyst Signature Settings (Active Session Only)</div>
        <div style="font-size: 0.9rem; opacity: 0.85;">Configure your credentials below. These remain saved for your current browser tab.</div>
    </div>
""", unsafe_allow_html=True)

col_user_name, col_user_title = st.columns(2)

with col_user_name:
    analyst_name = st.text_input(
        "Your Name (for email signature):",
        value="Anthony Cortez",
        key="session_analyst_name",
        help="Type your name once. It stays saved as long as this tab is open and resets when closed."
    )

with col_user_title:
    analyst_title = st.text_input(
        "Your Title:",
        value="Costco Benefits/HRIS Analyst",
        key="session_analyst_title",
        help="Type your job title once. It stays saved for this browser session."
    )

st.markdown("<br>", unsafe_allow_html=True)

# ==============================================================================
# INSTRUCTIONS GUIDE
# ==============================================================================
with st.expander("📖 Step-by-Step Instructions (How to Use This Tool)", expanded=True):
    st.markdown("""
    1. **Copy Unum Email Request:** Open the leave request email in your Benefits Analyst inbox and copy the full text (contains Employee Name, ID, and required date range).
    2. **Run SAP Time Report:**
       * Open SAP and go to **Cumulated Time Evaluation Results: Time Balances/Wage Types**.
       * Enter the **Personnel Number** (Employee ID).
       * Select **Other period** and enter the `Period` (Start Date) and `To` (End Date) specified in the Unum email.
       * Ensure Day balances is set to `0701` and Layout is `/TOTAL_HOURS`.
       * Execute the report (press **F8** or click the Green Checkmark).
    3. **Copy SAP Results to Clipboard:**
       * In SAP, click **System / File** $\rightarrow$ **Save to Local File** $\rightarrow$ select **In a clipboard** (or click the Clipboard icon).
    4. **Paste & Process:**
       * Paste **both** the copied Unum email text AND the copied SAP clipboard table together into the box below.
       * Click **🚀 Process Claim & Generate Report**.
    """)

# ==============================================================================
# ANONYMOUS SAMPLE DATA
# ==============================================================================
SAMPLE_ANONYMOUS_TEXT = """Employee Name: Anthony Cortez
Employee ID#: 01234567
Employee Phone #: (000) 000-0000
Leave Number: 99999999

Additional information is required to determine eligibility and/or available leave entitlement under STD/FMLA/State law for the above employee. Please reply to this e-mail by 3 p.m. EST on 08/05/26 with the requested information below. Failure to reply by the above time and date could delay Unum’s response to the employee’s request for leave.

Hours worked:

Total number of actual hours worked (including overtime hours) during the 12-month period of 08/02/25 through 08/05/26: ___________

------------------------------------------------------------------------------------------------------------------------
| Pers.No.|Employee/app.name  |Period|Date      |TmType|TimeTyText  |   Number|Cost Center  |PSubarea |Subarea|Cost Ctr|
------------------------------------------------------------------------------------------------------------------------
|  01234567|CORTEZ ANTHONY    |202508|08/20/2025|0701  |Hours Worked|    8.00 |STORE-101    |Retail   |0101   |1000101 |
|  01234567|CORTEZ ANTHONY    |202508|08/22/2025|0701  |Hours Worked|    8.00 |STORE-101    |Retail   |0101   |1000101 |
|  01234567|CORTEZ ANTHONY    |202607|07/20/2026|0701  |Hours Worked|    8.00 |STORE-101    |Retail   |0101   |1000101 |
------------------------------------------------------------------------------------------------------------------------"""

col_sample, col_clear = st.columns([3, 1])
with col_sample:
    if st.button("📋 Click Here to Insert Anonymous Example Claim"):
        st.session_state["claim_input"] = SAMPLE_ANONYMOUS_TEXT

with col_clear:
    if st.button("🗑️ Clear Text Box"):
        st.session_state["claim_input"] = ""

# Main Text Area
text_input = st.text_area(
    "Paste Combined Unum Request + SAP Clipboard Text Here:",
    key="claim_input",
    height=260,
    placeholder="Paste Unum Email and SAP Clipboard text here..."
)

# ==============================================================================
# AUTOMATION ENGINE
# ==============================================================================
def parse_sap_table(text):
    lines = text.split('\n')
    data_rows = []
    
    pipe_rows = [l.strip() for l in lines if l.strip().startswith('|') and not l.strip().startswith('|*') and 'Pers.No.' not in l]
    if pipe_rows:
        for line_str in pipe_rows:
            parts = [p.strip() for p in line_str.split('|')[1:-1]]
            if len(parts) >= 11:
                data_rows.append(parts[:11])
        headers = ['Pers.No.', 'Name', 'Period', 'Date', 'TmType', 'TimeTyText', 'Number', 'Cost Ctr', 'PSubarea', 'Subarea', 'Cost Ctr Ref']
        df = pd.DataFrame(data_rows, columns=headers)
    else:
        sap_match = re.search(r"(Pers\.No\..*)", text, re.DOTALL)
        if not sap_match:
            return None
        raw_sap = sap_match.group(1).strip()
        try:
            df = pd.read_csv(io.StringIO(raw_sap), sep=r'\t+|\s{2,}', engine='python')
        except Exception:
            df = pd.read_csv(io.StringIO(raw_sap), sep=None, engine='python')
        headers = [str(c).strip() for c in df.columns[:11]]
        df.columns = headers + list(df.columns[11:])
        df = df.dropna(subset=[headers[0]]).copy()

    df['Number'] = pd.to_numeric(df['Number'].astype(str).str.replace(',', ''), errors='coerce')
    return df

def process_combined_text(raw_text, sender_name, sender_title):
    emp_name_match = re.search(r"Employee Name:\s*(.+)", raw_text)
    emp_id_match = re.search(r"Employee ID#:\s*(\d+)", raw_text)
    leave_num_match = re.search(r"Leave Number:\s*(\d+)", raw_text)
    dates_match = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})\s+through\s+(\d{1,2}/\d{1,2}/\d{2,4})", raw_text)

    emp_name = emp_name_match.group(1).strip() if emp_name_match else "Employee"
    emp_id = emp_id_match.group(1).strip() if emp_id_match else "000000"
    leave_num = leave_num_match.group(1).strip() if leave_num_match else "N/A"

    start_date_str = dates_match.group(1) if dates_match else "08/02/2025"
    end_date_str = dates_match.group(2) if dates_match else "08/05/2026"

    clean_emp_id = str(emp_id).lstrip('0')
    last_name = emp_name.split()[-1]
    output_filename = f"{clean_emp_id} - {last_name}.xlsx"

    def parse_dt(d_str):
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try: return datetime.datetime.strptime(d_str, fmt).date()
            except ValueError: pass
        return datetime.date.today()

    req_start = parse_dt(start_date_str)
    req_end = parse_dt(end_date_str)

    df = parse_sap_table(raw_text)
    if df is None or df.empty:
        st.error("Could not parse SAP table. Ensure table rows start with '|' pipe symbols or standard tab formatting.")
        return None, None, None, []

    df['Date_dt'] = pd.to_datetime(df['Date']).dt.date
    df_filtered = df[(df['Date_dt'] >= req_start) & (df['Date_dt'] <= req_end)].copy()
    df_filtered = df_filtered.sort_values(by='Date_dt')

    # AUDIT CHECK: MISSING BOUNDARIES & >10 DAYS LEAVE GAPS
    missing_dates = []
    sap_warnings = []
    
    if req_start not in df_filtered['Date_dt'].values:
        missing_dates.append(req_start.strftime("%m/%d/%Y"))
    if req_end not in df_filtered['Date_dt'].values:
        missing_dates.append(req_end.strftime("%m/%d/%Y"))

    # Start-of-period gap check (> 10 Days)
    if not df_filtered.empty:
        first_sap_date = df_filtered['Date_dt'].min()
        start_gap = (first_sap_date - req_start).days
        if start_gap > 10:
            msg = f"Employee returned from leave on {first_sap_date.strftime('%m/%d/%Y')} (First worked date in SAP is {start_gap} days after requested start date {req_start.strftime('%m/%d/%Y')})."
            sap_warnings.append(msg)

        # End-of-period gap check (> 10 Days)
        last_sap_date = df_filtered['Date_dt'].max()
        end_gap = (req_end - last_sap_date).days
        if end_gap > 10:
            msg = f"Last worked date in SAP is {last_sap_date.strftime('%m/%d/%Y')} ({end_gap} days before requested end date {req_end.strftime('%m/%d/%Y')}). Check SAP Infotype 2001/2006 for leave/vacation."
            sap_warnings.append(msg)

    # Build Excel
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "FMLA Hours Log"
    ws_out.views.sheetView[0].showGridLines = True

    NAVY_PRIMARY = "0F172A"
    ZEBRA_FILL = PatternFill(start_color="F8FAFC", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", fill_type="solid")
    GOLD_TOTAL_FILL = PatternFill(start_color="FEF08A", fill_type="solid")
    CARD_BG_FILL = PatternFill(start_color="F1F5F9", fill_type="solid")
    ALERT_BG_FILL = PatternFill(start_color="FEF2F2", fill_type="solid")

    FONT_BANNER = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    FONT_HEADER = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    FONT_BODY = Font(name="Segoe UI", size=11, color="1E293B")
    FONT_BOLD = Font(name="Segoe UI", size=11, bold=True, color="0F172A")
    FONT_TOTAL = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
    FONT_KPI_VAL = Font(name="Segoe UI", size=26, bold=True, color="1E3A8A")
    FONT_KPI_LBL = Font(name="Segoe UI", size=9, bold=True, color="64748B")
    FONT_ALERT_TITLE = Font(name="Segoe UI", size=11, bold=True, color="991B1B")
    FONT_ALERT_BODY = Font(name="Segoe UI", size=11, bold=True, italic=True, color="B91C1C")

    BORDER_SUBTLE = Side(border_style="thin", color="E2E8F0")
    BORDER_GRID = Border(left=BORDER_SUBTLE, right=BORDER_SUBTLE, top=BORDER_SUBTLE, bottom=BORDER_SUBTLE)
    BORDER_TOTAL = Border(top=Side(border_style="medium", color="0F172A"), bottom=Side(border_style="double", color="0F172A"), left=BORDER_SUBTLE, right=BORDER_SUBTLE)
    BORDER_CARD = Border(left=Side(border_style="thick", color="2563EB"), right=BORDER_SUBTLE, top=BORDER_SUBTLE, bottom=BORDER_SUBTLE)

    ws_out.merge_cells("A1:K2")
    ws_out["A1"] = "  FMLA / STD HOURS WORKED VERIFICATION REPORT"
    ws_out["A1"].font = FONT_BANNER
    ws_out["A1"].fill = PatternFill(start_color=NAVY_PRIMARY, fill_type="solid")
    ws_out["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws_out["A3"] = "  Official Attendance & Shift Log | Benefits & HRIS Department"
    ws_out["A3"].font = Font(name="Segoe UI", size=10, italic=True, color="64748B")

    start_table_row = 5
    table_headers = ['Pers.No.', 'Name', 'Period', 'Date', 'TmType', 'TimeTyText', 'Number', 'Cost Ctr', 'PSubarea', 'Subarea', 'Cost Ctr Ref']
    for col_i, h_txt in enumerate(table_headers, start=1):
        cell = ws_out.cell(row=start_table_row, column=col_i, value=h_txt)
        cell.fill = PatternFill(start_color=NAVY_PRIMARY, fill_type="solid")
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_GRID

    ws_out.row_dimensions[start_table_row].height = 32

    filtered_vals = df_filtered.iloc[:, :11].values
    for idx, r_vals in enumerate(filtered_vals, start=start_table_row + 1):
        ws_out.row_dimensions[idx].height = 22
        row_fill = ZEBRA_FILL if idx % 2 == 0 else WHITE_FILL
        for col_i in range(1, 12):
            val = r_vals[col_i - 1]
            if col_i in [1, 3, 5, 10, 11]:
                try: val = int(float(val))
                except: pass
            elif col_i == 7:
                try: val = float(val)
                except: pass

            cell = ws_out.cell(row=idx, column=col_i, value=val)
            cell.font = FONT_BODY
            cell.fill = row_fill
            cell.border = BORDER_GRID

            if col_i in [1, 3, 5, 10, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_i == 4:
                cell.number_format = "MM/DD/YYYY"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_i == 7:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    total_row = start_table_row + len(filtered_vals) + 1
    ws_out.row_dimensions[total_row].height = 26
    ws_out.cell(row=total_row, column=6, value="Total Hours:").font = FONT_TOTAL
    ws_out.cell(row=total_row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws_out.cell(row=total_row, column=6).border = BORDER_TOTAL

    sum_cell = ws_out.cell(row=total_row, column=7, value=f"=SUM(G6:G{total_row-1})")
    sum_cell.font = FONT_TOTAL
    sum_cell.number_format = "#,##0.00"
    sum_cell.fill = GOLD_TOTAL_FILL
    sum_cell.alignment = Alignment(horizontal="right", vertical="center")
    sum_cell.border = BORDER_TOTAL

    for c in range(1, 12):
        if c not in [6, 7]:
            ws_out.cell(row=total_row, column=c).border = Border(top=Side(border_style="medium", color="0F172A"), bottom=Side(border_style="thin", color="0F172A"))

    # Missing Notes Under Table
    missing_note_row = (total_row - 1) + 5
    note_idx = 0
    for m_date in missing_dates:
        ws_out.cell(row=missing_note_row + note_idx, column=1, value=f"Employee did not work {m_date}").font = FONT_ALERT_BODY
        note_idx += 1

    for warn in sap_warnings:
        ws_out.cell(row=missing_note_row + note_idx, column=1, value=warn).font = FONT_ALERT_BODY
        note_idx += 1

    # Metrics Card
    ws_out.merge_cells("M1:P2")
    ws_out["M1"] = "CLAIM AUDIT & METRICS CARD"
    ws_out["M1"].font = FONT_BANNER
    ws_out["M1"].fill = PatternFill(start_color=NAVY_PRIMARY, fill_type="solid")
    ws_out["M1"].alignment = Alignment(horizontal="center", vertical="center")

    info_card = [
        ("Employee Name:", emp_name),
        ("Employee ID#:", str(clean_emp_id)),
        ("Leave Number:", str(leave_num)),
        ("Requested Period:", f"{req_start.strftime('%m/%d/%Y')} – {req_end.strftime('%m/%d/%Y')}"),
    ]

    for idx, (lbl, val) in enumerate(info_card, start=5):
        ws_out.cell(row=idx, column=13, value=lbl).font = Font(name="Segoe UI", size=10, bold=True, color="64748B")
        ws_out.merge_cells(start_row=idx, start_column=14, end_row=idx, end_column=16)
        ws_out.cell(row=idx, column=14, value=val).font = FONT_BOLD
        for c in range(13, 17):
            ws_out.cell(row=idx, column=c).fill = CARD_BG_FILL
            ws_out.cell(row=idx, column=c).border = BORDER_CARD

    ws_out.merge_cells("M11:N11")
    ws_out["M11"] = "TOTAL ACTUAL HOURS WORKED"
    ws_out["M11"].font = FONT_KPI_LBL
    ws_out["M11"].alignment = Alignment(horizontal="center")

    ws_out.merge_cells("M12:N14")
    ws_out["M12"] = f"=G{total_row}"
    ws_out["M12"].font = FONT_KPI_VAL
    ws_out["M12"].number_format = "#,##0.00"
    ws_out["M12"].alignment = Alignment(horizontal="center", vertical="center")

    ws_out.merge_cells("O11:P11")
    ws_out["O11"] = "TOTAL SHIFTS LOGGED"
    ws_out["O11"].font = FONT_KPI_LBL
    ws_out["O11"].alignment = Alignment(horizontal="center")

    ws_out.merge_cells("O12:P14")
    ws_out["O12"] = f"=COUNT(G6:G{total_row-1})"
    ws_out["O12"].font = FONT_KPI_VAL
    ws_out["O12"].alignment = Alignment(horizontal="center", vertical="center")

    for r in range(11, 15):
        for c in range(13, 17):
            cell = ws_out.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="F8FAFC", fill_type="solid")
            cell.border = BORDER_GRID

    ws_out.merge_cells("M16:P16")
    ws_out["M16"] = "MISSING DATES & AUDIT EXCEPTIONS"
    ws_out["M16"].font = FONT_ALERT_TITLE
    ws_out["M16"].fill = ALERT_BG_FILL
    ws_out["M16"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    all_exceptions = [f"Employee did not work {md}" for md in missing_dates] + sap_warnings
    if all_exceptions:
        for i, exc in enumerate(all_exceptions):
            r_idx = 17 + i
            ws_out.merge_cells(start_row=r_idx, start_column=13, end_row=r_idx, end_column=16)
            ws_out.cell(row=r_idx, column=13, value=exc).font = FONT_ALERT_BODY
    else:
        ws_out.merge_cells("M17:P17")
        ws_out["M17"] = "No missing boundary dates or extended gaps detected."
        ws_out["M17"].font = FONT_BOLD

    ws_out.column_dimensions['A'].width = 38
    ws_out.column_dimensions['B'].width = 28
    ws_out.column_dimensions['C'].width = 14
    ws_out.column_dimensions['D'].width = 16
    ws_out.column_dimensions['E'].width = 14
    ws_out.column_dimensions['F'].width = 17
    ws_out.column_dimensions['G'].width = 18
    ws_out.column_dimensions['H'].width = 18
    ws_out.column_dimensions['I'].width = 14
    ws_out.column_dimensions['J'].width = 14
    ws_out.column_dimensions['K'].width = 17
    ws_out.column_dimensions['L'].width = 4
    ws_out.column_dimensions['M'].width = 22
    ws_out.column_dimensions['N'].width = 22
    ws_out.column_dimensions['O'].width = 22
    ws_out.column_dimensions['P'].width = 28

    excel_buffer = io.BytesIO()
    wb_out.save(excel_buffer)
    excel_buffer.seek(0)

    # Dynamic Email Response
    total_hours_val = df_filtered['Number'].sum()
    
    email_notes = []
    if missing_dates:
        email_notes.append(f"did not work on {' and '.join(missing_dates)}")
    for warn in sap_warnings:
        email_notes.append(warn)

    notes_str = f" ({'; '.join(email_notes)})" if email_notes else ""

    email_response = f"""Per your request. Please see attached report to determine eligible hours. 

Employee logged {total_hours_val:,.2f} total hours worked{notes_str}. 

Direct any questions to your supervisor.

Thank you,

{sender_name}
{sender_title}"""

    return email_response, excel_buffer, output_filename, sap_warnings

# Process Action
if st.button("🚀 Process Claim & Generate Report"):
    if not text_input.strip():
        st.warning("Please paste text before clicking process.")
    else:
        email_txt, excel_data, filename, warnings = process_combined_text(
            text_input,
            analyst_name.strip() if analyst_name.strip() else "Anthony Cortez",
            analyst_title.strip() if analyst_title.strip() else "Costco Benefits/HRIS Analyst"
        )
        if email_txt:
            st.success("Claim audit complete!")

            # Display SAP Verification Warnings if gaps > 10 days exist
            if warnings:
                st.warning("⚠️ **SAP LEAVE DOUBLE-CHECK REQUIRED:**")
                for w in warnings:
                    st.write(f"• {w}")

            st.subheader("✉️ Ready-to-Send Email Reply")
            st.code(email_txt, language="text")
            
            st.subheader("📥 Executive Excel Report")
            st.download_button(
                label=f"Download {filename}",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
